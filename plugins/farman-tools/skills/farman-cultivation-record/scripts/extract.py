# -*- coding: utf-8 -*-
"""作業日誌(xlsx) から圃場ごとの作業レコードを抽出する。

    python extract.py --config ~/.farman/cultivation-record.json [-o records.json]

日誌の列位置は月ごとにずれることがあるため、**必ず見出し行から実測する**。
固定列番号で読むと数量を丸ごと取りこぼす（references/daily-log-structure.md 参照）。
"""
import argparse
import collections
import datetime
import json
import os
import re
import unicodedata

import openpyxl


# ---------------------------------------------------------------- 文字列処理
def norm(s):
    """全角英数・全角空白を半角へ。区切り記号を空白に統一する。"""
    s = unicodedata.normalize('NFKC', str(s))
    s = s.replace('、', ' ').replace('，', ' ').replace(',', ' ')
    s = re.sub(r'[.;:・/]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def to_num(v):
    """数量セル。'4000' のように文字列で入力されている行があるため数値化する。"""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r'[,\s，]', '', str(v))
    return float(t) if re.fullmatch(r'\d+(\.\d+)?', t) else None


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%m/%d')
    return str(v).strip()


def parse_md(v):
    """日付セル → (月, 日)。日付型と 'M/D' 文字列の両方に対応。"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.month, v.day
    m = re.match(r'\s*(\d{1,2})\s*/\s*(\d{1,2})', str(v))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


# ---------------------------------------------------------------- 圃場コード
class FieldParser:
    """圃場セルの文字列を台帳の圃場コードへ分解する。

    日誌は1セルに複数圃場をまとめて書く（`C.N.O.P.R` `A DM S C N O` `ＰＴＵＲ`）。
    区切りはドット・空白・全角空白・無しが混在し、全角英字も現れる。
    """

    def __init__(self, cfg, ledger_codes):
        self.codes = ledger_codes                      # {地区: set(コード)}
        self.aliases = {k.upper(): v for k, v in cfg.get('aliases', {}).items()}
        self.non_field = sorted(cfg.get('nonFieldLabels', []), key=len, reverse=True)
        self.districts = [d['ledger'] for d in cfg['districts']]
        self.log_names = sorted({d['logColumn'] for d in cfg['districts']}
                                | set(self.districts), key=len, reverse=True)
        self.fixed = {d['logColumn']: d.get('fixedCode') for d in cfg['districts']}

    def tokenize(self, text, district, fixed_code=None):
        """→ (確定コードのリスト, 未解釈の残り)"""
        s = norm(text)
        for name in self.log_names:                    # 「東井出C」等の地区名を除去
            s = s.replace(name, ' ')
        for w in self.non_field:                       # 圃場でない呼称を除去
            s = s.replace(norm(w), ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        if fixed_code:                                 # 「長坂→日野」のような固定コード
            return ([fixed_code], '') if s or fixed_code in text else ([], s)

        valid = self.codes.get(district, set())
        out, left, i = [], [], 0
        while i < len(s):
            if s[i] == ' ':
                i += 1
                continue
            two = s[i:i + 2].upper()
            if two in self.aliases:
                out.extend(self.aliases[two])
                i += 2
                continue
            if two in valid:                           # 1A・DM など2文字コードを優先
                out.append(two)
                i += 2
                continue
            one = s[i].upper()
            if one in valid:
                out.append(one)
                i += 1
                continue
            left.append(s[i])
            i += 1
        return out, ''.join(left).strip()

    def parse_other(self, text):
        """「その他」列の『長沢;S C K 東井出; B C』形式を地区ごとに分解する。"""
        s = norm(text)
        hits = sorted((s.find(d), d) for d in self.log_names if d in s)
        picked = []
        for pos, d in hits:
            if pos < 0 or any(p <= pos < p + len(dd) for p, dd in picked):
                continue                               # 「窪長沢」内の「長沢」を拾わない
            picked.append((pos, d))
        found = {}
        for k, (pos, d) in enumerate(picked):
            end = picked[k + 1][0] if k + 1 < len(picked) else len(s)
            dist = next((x['ledger'] for x in self._dcfg if x['logColumn'] == d
                         or x['ledger'] == d), d)
            codes, _ = self.tokenize(s[pos + len(d):end], dist, self.fixed.get(d))
            if codes:
                found.setdefault(dist, []).extend(codes)
        return found


# ---------------------------------------------------------------- 台帳
def read_ledger(cfg):
    """圃場台帳 → OrderedDict{圃場名: {area, struck}}
    取り消し線が引かれた行は「作成済み」を意味するため struck=True を立てる。"""
    wb = openpyxl.load_workbook(cfg['paths']['ledger'])
    ws = wb[cfg['paths'].get('ledgerSheet') or wb.sheetnames[0]]
    col = {}
    for r in range(1, 8):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() in ('圃場名', '面積'):
                col[str(v).strip()] = c
    c_name, c_area = col.get('圃場名', 3), col.get('面積', 5)

    groups, cur = [], None
    for r in range(1, ws.max_row + 1):
        name = ws.cell(row=r, column=c_name).value
        if name and str(name).strip() in ('圃場名',):
            continue
        area = ws.cell(row=r, column=c_area).value
        struck = any(bool(ws.cell(row=r, column=c).font and ws.cell(row=r, column=c).font.strike)
                     for c in (c_name, c_name + 1))
        if name:
            cur = dict(name=norm(name).replace(' ', ''), area=0.0, struck=struck)
            groups.append(cur)
        if cur is not None and isinstance(area, (int, float)):
            cur['area'] += float(area)

    merged = collections.OrderedDict()
    for g in groups:                                   # 同名ブロックが分かれている場合は合算
        if g['name'] in merged:
            merged[g['name']]['area'] += g['area']
            merged[g['name']]['struck'] |= g['struck']
        else:
            merged[g['name']] = g

    for a, b, new in cfg.get('mergedFields', []):      # 常に一体記録される圃場を統合
        if a in merged and b in merged:
            out = collections.OrderedDict()
            for k, v in merged.items():
                if k == a:
                    out[new] = dict(name=new, area=merged[a]['area'] + merged[b]['area'],
                                    struck=merged[a]['struck'] or merged[b]['struck'])
                elif k != b:
                    out[k] = v
            merged = out
    return merged


def split_name(cfg, name):
    """台帳名 → (地区, コード)"""
    for d in sorted(cfg['districts'], key=lambda x: -len(x['ledger'])):
        if name.startswith(d['ledger']):
            return d['ledger'], (d.get('fixedCode') or name[len(d['ledger']):])
    return None, None


# ---------------------------------------------------------------- 日誌の列解決
def resolve_cols(ws, cfg):
    """見出し行から列位置を実測する。月によって作物以降が1列ずれることがあるため、
    固定値を使うと定植・播種の数量を丸ごと落とす。"""
    h = {}
    for r in (1, 2):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v:
                h.setdefault(str(v).strip(), c)
    if '圃場' not in h:
        return None
    base = h['圃場']
    # 年によって圃場列の顔ぶれが違う(古い月は「長沢/窪長沢/その他」だけ)。
    # 見出しに実在する列だけを採用し、無い地区は黙って飛ばす。
    cols = {}
    for d in cfg['districts']:
        c = h.get(d['logColumn'])
        if c and c >= base:                            # 作業者欄の「その他」等と衝突させない
            cols[c] = d['logColumn']
    other = (max(cols) + 1) if cols else None
    juryo, teishoku = h.get('重量情報'), h.get('定植・播種情報')
    naiyo = h.get('内容（栽培関係）') or h.get('作業内容')
    return dict(
        hojo=cols, other=other,
        naiyo=naiyo, naiyo2=h.get('内容（その他）'),
        crop=h.get('作物') or h.get('品目'), variety=h.get('品種'),
        mach=h.get('使用機械'), mat=(h.get('資材') or h.get('使用資材')),
        biko=h.get('備考'),
        qkg=(juryo + 2) if juryo else None,            # 重量情報: kg及びℓ/個数/合計kg数
        qty=(teishoku + 3) if teishoku else None,      # 定植・播種情報: ○穴/粒/枚数/合計粒
    )


# ---------------------------------------------------------------- 抽出本体
def run(cfg):
    wb = openpyxl.load_workbook(cfg['paths']['dailyLog'], data_only=True)
    ledger = read_ledger(cfg)
    codes = collections.defaultdict(set)
    for nm in ledger:
        d, c = split_name(cfg, nm)
        if d:
            codes[d].add(c)

    fp = FieldParser(cfg, codes)
    fp._dcfg = cfg['districts']
    log2ledger = {d['logColumn']: d['ledger'] for d in cfg['districts']}

    records = collections.defaultdict(list)
    unparsed = collections.Counter()
    prefix = cfg['year']['logSheetPrefix']

    # タブは新しい月が左に並んでいることがあるため、必ず月順に処理する
    # (順序が変わると肥料・機械・作物の一覧の並びが変わってしまう)
    months = sorted((n for n in wb.sheetnames if re.fullmatch(re.escape(prefix) + r'\d+', n)),
                    key=lambda n: int(n[len(prefix):]))
    for name in months:
        ws = wb[name]
        C = resolve_cols(ws, cfg)
        if not C or not C['naiyo']:
            continue
        for r in range(2, ws.max_row + 1):
            mm, dd = parse_md(ws.cell(row=r, column=1).value)
            if mm is None:
                continue
            found = collections.defaultdict(list)
            for col, logname in C['hojo'].items():
                v = cell(ws, r, col)
                if not v:
                    continue
                dist = log2ledger.get(logname, logname)
                got, left = fp.tokenize(v, dist, fp.fixed.get(logname))
                found[dist].extend(got)
                if left:
                    unparsed['%s|%s' % (dist, v)] += 1
            ov = cell(ws, r, C['other']) if C['other'] else ''
            if ov:
                for d, cs in fp.parse_other(ov).items():
                    found[log2ledger.get(d, d)].extend(cs)
            if not found:
                continue

            w1 = cell(ws, r, C['naiyo'])
            w2 = cell(ws, r, C['naiyo2']) if C['naiyo2'] else ''
            rec = dict(
                m=mm, d=dd, work=(w1 or w2), work2=(w2 if w1 else ''),
                crop=cell(ws, r, C['crop']) if C['crop'] else '',
                variety=cell(ws, r, C['variety']) if C['variety'] else '',
                mach=cell(ws, r, C['mach']) if C['mach'] else '',
                mat=cell(ws, r, C['mat']) if C['mat'] else '',
                qty=to_num(ws.cell(row=r, column=C['qty']).value) if C['qty'] else None,
                qkg=to_num(ws.cell(row=r, column=C['qkg']).value) if C['qkg'] else None,
                note=cell(ws, r, C['biko']) if C['biko'] else '')
            for dist, cs in found.items():
                for code in dict.fromkeys(cs):
                    records[(dist, code)].append(rec)

    # ---- 委託タブ(指定圃場のみ)
    dele = (cfg.get('options') or {}).get('delegatedTab') or {}
    if dele.get('sheet') in wb.sheetnames and dele.get('onlyField'):
        dist, code = dele['onlyField'].split('|')
        ws = wb[dele['sheet']]
        C = resolve_cols(ws, cfg)
        if C:
            hojo_col = next((c for c, n in C['hojo'].items()
                             if log2ledger.get(n, n) == dist), None)
            for r in range(2, ws.max_row + 1):
                mm, dd = parse_md(ws.cell(row=r, column=1).value)
                if mm is None or hojo_col is None:
                    continue
                if cell(ws, r, hojo_col).strip().upper() != code.upper():
                    continue
                records[(dist, code)].append(dict(
                    m=mm, d=dd, work=cell(ws, r, C['naiyo']), work2='',
                    crop='', variety='', mach='',
                    mat=cell(ws, r, C['mat']) if C['mat'] else '',
                    qty=None, qkg=None, note='(委託作業)'))

    return ledger, records, unparsed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--config', required=True)
    ap.add_argument('-o', '--out', default='records.json')
    # Cowork等、設定ファイルのパスが使えない環境向けの上書き
    ap.add_argument('--daily-log', help='作業日誌xlsxのパス(設定より優先)')
    ap.add_argument('--ledger', help='圃場台帳xlsxのパス(設定より優先)')
    a = ap.parse_args()
    cfg = json.load(open(os.path.expanduser(a.config), encoding='utf-8'))
    if a.daily_log:
        cfg['paths']['dailyLog'] = a.daily_log
    if a.ledger:
        cfg['paths']['ledger'] = a.ledger

    ledger, recs, unp = run(cfg)
    if unp:
        print('--- 未解釈が残った圃場セル(要確認) ---')
        for k, n in unp.most_common(20):
            print('  %3d  %s' % (n, k))
    print('--- 圃場別 レコード数 ---')
    for name, g in ledger.items():
        d, c = split_name(cfg, name)
        mark = ' [作成済み]' if g['struck'] else ''
        print('  %-14s %4d%s' % (name, len(recs.get((d, c), [])), mark))
    json.dump({'%s|%s' % k: v for k, v in recs.items()},
              open(a.out, 'w', encoding='utf-8'), ensure_ascii=False)
    print('\nwrote %s' % a.out)


if __name__ == '__main__':
    main()
