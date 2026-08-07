# -*- coding: utf-8 -*-
"""抽出レコード → 地名ごとの栽培履歴ブック(xlsx)を生成する。

    python build.py --config ~/.farman/cultivation-record.json [-r records.json] [--only 長沢]

1圃場=1シート。ひな形の書式を保ったまま、ヘッダー部と作業履歴を差し替える。
耕作履歴の無い圃場も「休耕・不作付」のシートを作る（提出上の欠落を防ぐため）。
"""
import argparse
import collections
import copy
import datetime
import json
import os
import re
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.properties import PageSetupProperties

import extract as EX

THIN, MED = Side(style='thin'), Side(style='medium')
UNKNOWN_FILL = PatternFill('solid', fgColor='FFF2A8')   # 数量が日報に無いセル
LABEL_FONT = Font(name='ＭＳ Ｐゴシック', size=10, bold=True)
BODY_FONT = Font(name='ＭＳ Ｐゴシック', size=8)
HDR_FONT = Font(name='ＭＳ Ｐゴシック', size=11)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

UNIT = r'(倍|袋|体|反|ℓ|L|kg|㎏|g|t|立米|c/s|粒|本|枚|台|車|回|m|M|リットル)'
NOT_FERT_HINT = ('から', 'くらい', 'まで', '終了', '残り', '南側', '北側', '東側', '西側', '側')


def norm(s):
    return unicodedata.normalize('NFKC', str(s)).strip()


# ---------------------------------------------------------------- 肥料・資材
def fert_names(text, drop):
    """施肥・追肥の記述から肥料名だけを取り出す(数量・希釈倍率は落とす)。"""
    t = re.sub(r'^(施肥|追肥|元肥)', ' ', norm(text))
    t = re.sub(r'[（(]([^)）]*)[)）]', r' \1 ', t)
    out = []
    for p in re.split(r'[ 　、,／/]+', t):
        p = p.strip(' :：0123456789.')
        if not p:
            continue
        p = re.sub(r'[:：]?\s*[\d.,]+\s*' + UNIT + r'?$', '', p).strip()
        if not p or re.fullmatch(r'[\d.,]+', p) or len(p) == 1:
            continue
        if any(w in p for w in NOT_FERT_HINT) or p in drop:
            continue
        out.append(p)
    return out


def summarize(recs, cfg):
    canon = cfg.get('fertilizerCanon', {})
    drop = set(cfg.get('fertilizerDrop', []))
    prot_map = cfg.get('protectants', {})
    mat2mach = cfg.get('materialToMachine', {})
    ferts, prot, mats, machs = [], [], [], []
    for r in recs:
        work = r.get('work') or ''
        w = work + ' ' + (r.get('work2') or '')
        if any(k in w for k in ('施肥', '追肥', '元肥', '堆肥')):
            for nm in fert_names(w, drop):
                (prot if nm in prot_map else ferts).append(prot_map.get(nm, canon.get(nm, nm)))
        for key, val in prot_map.items():
            if key in w:
                prot.append(val)
        mat = norm(r.get('mat') or '')
        if mat:
            # 読み替えは drop より先に見る(機械名が drop に入っていても機械へ回すため)
            if mat in mat2mach:                       # 資材欄に書かれた機械名を機械へ回す
                machs.append(mat2mach[mat])
            elif any(k in work for k in ('施肥', '追肥')):
                ferts += [canon.get(x, x) for x in fert_names(mat, drop)]
            elif mat not in drop:
                mats.append(mat)
        if r.get('mach'):
            machs.append(norm(r['mach']))
        if 'マルチ' in w:
            mats.append('マルチ')
    u = lambda xs: list(dict.fromkeys(x for x in xs if x))
    return dict(fert=u(ferts), prot=u(prot), mat=u(mats), mach=u(machs))


def clean_work(text, cfg):
    """作業内容から数量・希釈倍率を落とす。肥料名は残す(種類は記録が必要なため)。"""
    canon = cfg.get('fertilizerCanon', {})
    drop = set(cfg.get('fertilizerDrop', []))
    t = norm(text)
    for key, val in cfg.get('protectants', {}).items():
        if key in t:
            return val + '散布'
    t = re.sub(r'[\d.,]+\s*' + UNIT, ' ', t)
    t = re.sub(r'[:：]\s*', ' ', t)
    t = re.sub(r'(?<=[^\d])[\d.,]+(?=\s|$|[)）])', ' ', t)
    t = re.sub(r'[（(]\s*[)）]', '', t)
    t = re.sub(r'\s+', ' ', t).strip(' 、,')
    m = re.match(r'^(.*?)(施肥|追肥|元肥)', t)
    if m:
        names = [canon.get(x, x) for x in fert_names(t, drop)]
        names = [n for n in dict.fromkeys(names) if n not in ('耕耘', 'マルチ張り')]
        head = m.group(1).strip(' 、,')
        body = m.group(2) + ('(%s)' % '、'.join(names) if names else '')
        return (head + '、' + body) if head else body
    return t


def dedup_log(recs, cfg, dropped):
    """(月日, 作物/品種, 作業内容) で重複を除き日付順に並べる。事務・出荷作業は除外。"""
    non_cult = tuple(cfg.get('nonCultivation', []))
    seen, out = set(), []
    for r in recs:
        work = (r.get('work') or '').strip()
        extra = (r.get('work2') or '').strip()
        if extra and extra != work:
            work = '%s(%s)' % (work, extra)
        work = clean_work(work, cfg)[:40]
        crop = norm(r.get('crop') or '')
        if crop and r.get('variety'):
            crop = '%s/%s' % (crop, norm(r['variety']))
        crop = crop[:24]
        key = (r['m'], r['d'], crop, work)
        if key in seen or not work:
            continue
        if any(w in work for w in non_cult):
            dropped[work] += 1
            continue
        seen.add(key)
        out.append(dict(m=r['m'], d=r['d'], crop=crop, work=work,
                        itaku='(委託作業)' in (r.get('note') or '')))
    out.sort(key=lambda x: (x['m'], x['d']))
    return out


def is_fallow(log, cfg):
    words = tuple(cfg.get('nonFarming', []))
    return (not log) or all(any(w in e['work'] for w in words) for e in log)


def seed_combos(recs):
    """定植・播種の記録 → (作物/品種, 使用数)。数量が日報に無いものは None。"""
    d = collections.OrderedDict()
    for r in recs:
        w = (r.get('work') or '') + (r.get('work2') or '')
        crop = norm(r.get('crop') or '')
        if not crop:
            continue
        planted = any(k in w for k in ('定植', '播種'))
        key = crop + ('/' + norm(r['variety']) if r.get('variety') else '')
        e = d.setdefault(key, dict(q=0, kinds=set(), planted=False))
        e['planted'] |= planted
        if planted:
            e['kinds'].add('播種' if '播種' in w else '定植')
            if r.get('qty'):
                e['q'] += r['qty']
    items = [(k, v) for k, v in d.items() if v['planted']] or list(d.items())
    with_var = {k.split('/')[0] for k, _ in items if '/' in k}
    items = [(k, v) for k, v in items if '/' in k or k not in with_var]
    out = []
    for k, v in items:
        unit = '粒' if v['kinds'] == {'播種'} else ('本' if v['kinds'] == {'定植'} else '本・粒')
        out.append((k, '{:,}{}'.format(int(v['q']), unit) if v['q'] else None))
    return out


# ---------------------------------------------------------------- シート描画
def box(ws, row, c1, c2, left=THIN, right=MED, top=MED, bottom=MED):
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = Border(
            left=left if c == c1 else None, right=right if c == c2 else None,
            top=top, bottom=bottom)


def put_label(ws, row, text):
    c = ws.cell(row=row, column=5)
    c.value, c.font, c.alignment = text, copy.copy(LABEL_FONT), copy.copy(CTR)
    c.border = Border(left=MED, right=THIN, top=MED, bottom=MED)


def fill_sheet(ws, cfg, name, area, log, summ, fallow, seeds):
    for rng in [str(r) for r in ws.merged_cells.ranges]:
        if int(re.search(r'(\d+)', rng).group(1)) >= 6:
            ws.unmerge_cells(rng)
    for r in range(6, ws.max_row + 1):
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
            cell.border, cell.fill = Border(), PatternFill()
        ws.row_dimensions[r].height = None

    hdr = cfg['header']
    ws['F2'] = name
    ws['F3'] = cfg['year']['label']
    made = hdr.get('madeOn')
    ws['J4'] = (datetime.datetime.strptime(made, '%Y-%m-%d') if made
                else datetime.datetime.now().replace(
                    hour=0, minute=0, second=0, microsecond=0))   # 省略時は実行日

    for r, lab, val in ((6, '圃場名', '%s(%s)' % (name, area)),
                        (7, '栽培面積', area), (8, '責任者', hdr['staff'])):
        ws.cell(row=r, column=1).value = lab
        ws.cell(row=r, column=1).font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2)
        c.value, c.font = val, Font(name='ＭＳ Ｐゴシック', size=11)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 20.25

    ws.merge_cells('A10:D11')
    a = ws['A10']
    a.value = 'アルバイト・体験参加者には有機ＪＡＳの取り組みについて周知を行っています。'
    a.font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True)
    a.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 苗・種の使用数: 作物/品種ごと。5組ずつ折り返す可変ブロック
    PER, r0 = 5, 6
    bands = max(1, -(-len(seeds) // PER))
    for b in range(bands):
        rn_, rq = r0 + 2 * b, r0 + 2 * b + 1
        chunk = seeds[b * PER:(b + 1) * PER]
        box(ws, rn_, 6, 10, top=(MED if b == 0 else THIN), bottom=THIN)
        box(ws, rq, 6, 10, top=THIN, bottom=(MED if b == bands - 1 else THIN))
        for i in range(PER):
            cn, cq = ws.cell(row=rn_, column=6 + i), ws.cell(row=rq, column=6 + i)
            cn.font = cq.font = copy.copy(BODY_FONT)
            cn.alignment = cq.alignment = copy.copy(CTR)
            if i < len(chunk):
                cn.value, qty = chunk[i][0], chunk[i][1]
                if qty:
                    cq.value = qty
                else:
                    cq.fill = UNKNOWN_FILL      # 数量不明 → 塗って明示
        ws.row_dimensions[rn_].height = 26
        ws.row_dimensions[rq].height = 16
    ws.merge_cells(start_row=r0, start_column=5, end_row=r0 + 2 * bands - 1, end_column=5)
    put_label(ws, r0, '苗・種の使用数')

    rf = r0 + 2 * bands
    put_label(ws, rf, '肥料の使用量')
    ws.merge_cells(start_row=rf, start_column=6, end_row=rf, end_column=10)
    ws.cell(row=rf, column=6).value = '、'.join(summ['fert']) if summ['fert'] else '―'
    ws.cell(row=rf, column=6).alignment = copy.copy(LEFT)
    ws.cell(row=rf, column=6).font = copy.copy(BODY_FONT)
    box(ws, rf, 6, 10)
    ws.row_dimensions[rf].height = 30

    for row, lab, vals in ((rf + 1, '使用資材', summ['prot'] + summ['mat']),
                           (rf + 2, '使用機械', summ['mach'])):
        put_label(ws, row, lab)
        box(ws, row, 6, 10)
        for i in range(5):
            c = ws.cell(row=row, column=6 + i)
            c.font, c.alignment = copy.copy(BODY_FONT), copy.copy(CTR)
            if i == 0 and not vals:
                c.value = '―'
            elif i < len(vals):
                c.value = '、'.join(vals[4:]) if (i == 4 and len(vals) > 5) else vals[i]
        ws.row_dimensions[row].height = 22 if lab == '使用資材' else 30
    rm = rf + 2

    notes = []
    if fallow:
        notes.append('※当年度は休耕(不作付)のため、栽培作業実績はありません。')
    if any(e['itaku'] for e in log):
        notes.append('※委託作業分を含む。')
    rn = rm
    if notes:
        rn = rm + 1
        ws.merge_cells(start_row=rn, start_column=5, end_row=rn, end_column=10)
        ws.cell(row=rn, column=5).value = ' '.join(notes)
        ws.cell(row=rn, column=5).alignment = copy.copy(LEFT)
        ws.cell(row=rn, column=5).font = copy.copy(BODY_FONT)
        box(ws, rn, 5, 10, left=MED)
        ws.row_dimensions[rn].height = 22

    # 作業履歴: 月日/作物・品種/作業内容 を2ブロック並べる
    ws.column_dimensions['B'].width = 19
    ws.column_dimensions['G'].width = 19
    top = max(rn, 11) + 2
    entries = [] if fallow else log
    half = max(1, -(-len(entries) // 2))
    last_row = top + max(half, 12)

    for r in range(top, last_row + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.row_dimensions[r].height = 14.25
    for dcol, ccol, wcol, items in ((1, 2, 3, entries[:half]), (6, 7, 8, entries[half:])):
        for col, lab in ((dcol, '月日'), (ccol, '作物/品種'), (wcol, '作業内容')):
            h = ws.cell(row=top, column=col)
            h.value, h.font, h.alignment = lab, copy.copy(HDR_FONT), copy.copy(CTR)
        for i, e in enumerate(items):
            r = top + 1 + i
            dc = ws.cell(row=r, column=dcol)
            dc.value = datetime.datetime(cfg['year']['calendar'], e['m'], e['d'])
            dc.number_format = 'm/d;@'
            dc.font, dc.alignment = copy.copy(BODY_FONT), copy.copy(CTR)
            cc = ws.cell(row=r, column=ccol)
            cc.value, cc.font, cc.alignment = e['crop'], copy.copy(BODY_FONT), copy.copy(CTR)
            wc = ws.cell(row=r, column=wcol)
            wc.value, wc.font, wc.alignment = e['work'], copy.copy(BODY_FONT), copy.copy(LEFT)
    if fallow:
        ws.cell(row=top + 1, column=1).value = '―'
        ws.cell(row=top + 1, column=1).alignment = copy.copy(CTR)
        ws.cell(row=top + 1, column=3).value = '当年度は休耕(不作付)。栽培作業実績なし。'
        ws.cell(row=top + 1, column=3).font = copy.copy(BODY_FONT)
    for r in range(top, last_row + 1):
        for a_, b_ in ((3, 5), (8, 10)):
            ws.merge_cells(start_row=r, start_column=a_, end_row=r, end_column=b_)

    ws.print_area = 'A1:J%d' % last_row
    ws.print_title_rows = '1:%d' % top
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4


def district_of(cfg, name):
    for d in sorted(cfg['districts'], key=lambda x: -len(x['ledger'])):
        if name.startswith(d['ledger']):
            return d['ledger']
    return 'その他'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--config', required=True)
    ap.add_argument('-r', '--records', default='records.json')
    ap.add_argument('--only', help='この地名のブックだけ作り直す(他ファイルに触れない)')
    # Cowork等、設定ファイルのパスが使えない環境向けの上書き
    ap.add_argument('--ledger', help='圃場台帳xlsxのパス(設定より優先)')
    ap.add_argument('--template', help='ひな形xlsxのパス(設定より優先)')
    ap.add_argument('--out-dir', help='出力先フォルダ(設定より優先)')
    a = ap.parse_args()
    cfg = json.load(open(os.path.expanduser(a.config), encoding='utf-8'))
    cfg.setdefault('paths', {})          # 設定にパスが無くても引数で補える
    for key, val in (('ledger', a.ledger), ('template', a.template), ('outputDir', a.out_dir)):
        if val:
            cfg['paths'][key] = val
    for key in ('ledger', 'template', 'outputDir'):
        if not cfg['paths'].get(key):
            ap.error('%s のパスが設定にも引数にもありません' % key)
    recs = json.load(open(a.records, encoding='utf-8'))
    ledger = EX.read_ledger(cfg)
    exclude = set(cfg.get('excludeFields', []))

    by_dist = collections.OrderedDict()
    for g in ledger.values():
        if g['struck'] or g['name'] in exclude:        # 取り消し線=作成済み
            continue
        by_dist.setdefault(district_of(cfg, g['name']), []).append(g)

    outdir = os.path.expanduser(cfg['paths']['outputDir'])
    os.makedirs(outdir, exist_ok=True)
    dropped, report = collections.Counter(), []

    for dist, fields in by_dist.items():
        if a.only and dist != a.only:
            continue
        wb = openpyxl.load_workbook(cfg['paths']['template'])
        base = wb[cfg['paths'].get('templateSheet') or wb.sheetnames[0]]
        for s in list(wb.sheetnames):
            if s != base.title:
                del wb[s]
        made = [(wb.copy_worksheet(base), g) for g in fields]
        for ws, g in made:
            ws.title = g['name'][:31]
        del wb[base.title]

        for ws, g in made:
            d, c = EX.split_name(cfg, g['name'])
            raw = recs.get('%s|%s' % (d, c), [])
            log = dedup_log(raw, cfg, dropped)
            fallow = is_fallow(log, cfg)
            seeds = [] if fallow else seed_combos(raw)
            area = ('%ga' % round(g['area'], 2)) if g['area'] else '面積要確認'
            fill_sheet(ws, cfg, g['name'], area, log, summarize(raw, cfg), fallow, seeds)
            report.append((g['name'], len(log), len(seeds),
                           sum(1 for _, q in seeds if not q), fallow))

        out = os.path.join(outdir, '%s_栽培履歴_%s.xlsx' % (cfg['year']['calendar'], dist))
        wb.save(out)
        print('saved %-50s %2dシート' % (out, len(made)))

    print('\n%-14s %6s %6s %8s %s' % ('圃場', '作業', '苗種', '数量不明', '状態'))
    for name, n, ns, nu, fallow in report:
        print('%-14s %6d %6d %8d %s' % (name, n, ns, nu, '休耕・不作付' if fallow else ''))
    if dropped:
        print('\n--- 履歴から除外した事務・出荷作業 (計%d件) ---' % sum(dropped.values()))
        for w, n in dropped.most_common(15):
            print('  %3d  %s' % (n, w))
    print('\n計 %d シート / 休耕 %d' % (len(report), sum(1 for r in report if r[4])))


if __name__ == '__main__':
    main()
